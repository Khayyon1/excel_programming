from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet("Dish Raw Data")

# Headers
ws['A1'] = 'Target'
ws['D1'] = 'Baseline'
ws['H1'] = 'Percentage'
ws['I1'] = 'Target'
ws['J1'] = 'Baseline'
ws['L1'] = 'Adults'
ws['M1'] = 'Overindex'

# Target groups
groups = ['a18_24', 'a25_34', 'a35_44', 'a45_54','a55_64','a65_74', 'a75', 'age_total',
          'kids_0_2', 'kids_3_5','kids_6_10', 'kids_11_15', 'kids_16_17', 'kids_no', 'kids_total',
          'male', 'female', 'gender_total', 'married', 'single','marital_total',
          'hhi_50','hhi_51_75', 'hhi_76_100','hhi_101_125', 'hhi_126_150', 'hhi_151p','income_total',
          'african_american','asian','caucasian','hispanic', 'other_ethnicity', 'ethnic_total', 'highschool_or_less',
          'some_college','bachelor_degree','graduate_degree', 'education_total']

# This is A1 Groups
for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, max_row=40)):
    for cell in row:
        cell.value = groups[idx]

# This is Baseline Groups
for idx, row in enumerate(ws.iter_rows(min_row=2, min_col=4, max_col=4, max_row=40)):
    for cell in row:
        cell.value = groups[idx]

# This is for Percentage
for idx, row in enumerate(ws.iter_rows(min_row=2, min_col=8, max_col=8, max_row=40)):
    for cell in row:
        name = groups[idx]
        if 'total' not in name:
            cell.value = name
            
# This is for Column by Overindex
indices = ['Kids','Gender', 'Marital Status', 'Income', 'Ethnicity','Education']
for idx, row in enumerate(ws.iter_rows(min_row=2, min_col=12, max_col=12, max_row=39)):
    for cell in row:
        name = groups[idx]
        if 'total' not in name:
            cell.value = name
        else:
            print("Hi")
            cell.value = indices.pop(0)
        print(cell.value)

wb.save('Demographic_Breakdown_Aldi_RawData.xlsx')